import pandas as pd
import os
import time
import bleach
import pytz 
import requests 
from typing import Optional, List, Dict, Any, Union, cast 
from datetime import datetime, timedelta
from io import BytesIO
from collections import Counter

from fastapi import FastAPI, HTTPException, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# --- 0. KONFIGURASI AWAL ---
load_dotenv()

app = FastAPI(
    title="SiJAGAD API",
    description="Sistem Monitoring & Arsip Digital",
    version="1.0.0"
)

# --- 1. SETUP DATABASE & TELEGRAM ---
SUPABASE_URL: str = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY: str = os.getenv("SUPABASE_SERVICE_KEY", "")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("⚠️  WARNING: SUPABASE_URL atau SUPABASE_SERVICE_KEY belum diset di .env")

try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as e:
    print(f"❌ Gagal koneksi ke Supabase: {e}")


# --- 2. MODELS ---
class LetterSchema(BaseModel):
    vendor: str
    pekerjaan: str
    nomor_kontrak: str
    tanggal_awal_kontrak: str
    nominal_jaminan: int
    jenis_garansi: str
    nomor_garansi: str
    bank_penerbit: str
    tanggal_awal_garansi: str
    tanggal_akhir_garansi: str
    status: str
    kategori: str
    file_url: Optional[str] = None
    user_email: Optional[str] = "System"
    lokasi: Optional[str] = None


# --- 3. HELPER FUNCTIONS ---
def sanitize_text(text: str) -> str:
    if not text: return ""
    return bleach.clean(text, tags=[], strip=True)

def log_activity_bg(user_email: str, action: str, target: str):
    try:
        supabase.table("activity_sijagad").insert({
            "user_email": user_email,
            "action": action,
            "target": target,
            "created_at": datetime.now().isoformat()
        }).execute()
        print(f"📝 [Log] {action}: {target}")
    except Exception as e:
        print(f"❌ [Log Error]: {e}")

def send_telegram_notif(message: str, specific_chat_id: Optional[str] = None):
    """Kirim pesan ke Telegram (Bisa Broadcast ke Grup Default atau Balas Chat Tertentu)"""
    target_chat_id = specific_chat_id or TELEGRAM_CHAT_ID
    
    if not TELEGRAM_BOT_TOKEN or not target_chat_id:
        print("⚠️ Telegram Config Missing")
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {"chat_id": target_chat_id, "text": message, "parse_mode": "Markdown"}
        requests.post(url, json=payload)
    except Exception as e:
        print(f"❌ Gagal kirim Telegram: {e}")

def generate_upcoming_report_text() -> str:
    """Helper Function: Membuat teks laporan H-90"""
    try:
        tz_manado = pytz.timezone('Asia/Makassar') 
        today = datetime.now(tz_manado).date()
        future_date = (today + timedelta(days=90)).strftime('%Y-%m-%d')
        today_str = today.strftime('%Y-%m-%d')

        response = supabase.table("letters").select("vendor, nomor_kontrak, tanggal_akhir_garansi") \
            .eq("is_deleted", False) \
            .gte("tanggal_akhir_garansi", today_str) \
            .lte("tanggal_akhir_garansi", future_date) \
            .neq("status", "Expired") \
            .neq("status", "Selesai") \
            .order("tanggal_akhir_garansi", desc=False) \
            .execute()
            
        letters = cast(List[Dict[str, Any]], response.data or [])
        
        if not letters:
            return "✅ *AMAN TERKENDALI*\nTidak ada surat yang akan expired dalam 90 hari ke depan."

        report_lines = []
        for item in letters:
            try:
                tgl_akhir_str = str(item.get('tanggal_akhir_garansi'))
                tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d').date()
                sisa_hari = (tgl_akhir - today).days
                vendor = item.get('vendor', 'Unknown')
                kontrak = item.get('nomor_kontrak', '-')

                if sisa_hari <= 7: icon = "🔥" 
                elif sisa_hari <= 30: icon = "⚠️"
                else: icon = "⏳"

                report_lines.append(f"{icon} *{vendor}*\n   └ ⏰ Sisa: *{sisa_hari} Hari* ({tgl_akhir_str})\n   └ 📄 No: `{kontrak}`")
            except: continue

        display_lines = report_lines[:15]
        header = f"📊 *UPDATE SISA WAKTU SURAT* 📊\n_Per Tanggal: {today_str}_\n\n"
        content = "\n".join(display_lines)
        footer = f"\n\nTotal: {len(letters)} Surat mendekati jatuh tempo."
        if len(letters) > 15: footer += f"\n_(...dan {len(letters)-15} lainnya)_"

        return header + content + footer
    except Exception as e:
        return f"❌ Terjadi kesalahan sistem: {str(e)}"


# --- 4. MIDDLEWARE ---
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        return response

app.add_middleware(SecurityHeadersMiddleware)

@app.middleware("http")
async def add_process_time_header(request: Request, call_next):
    response = await call_next(request)
    return response

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


# --- 5. ENDPOINTS UTAMA ---

@app.get("/")
def read_root():
    return {"status": "online", "system": "SiJAGAD API Ready 🚀"}

# 🔥 ENDPOINT BARU: MENERIMA PESAN DARI TELEGRAM (WEBHOOK)
@app.post("/telegram-webhook")
async def telegram_webhook(request: Request, background_tasks: BackgroundTasks):
    try:
        data = await request.json()
        
        # Cek apakah ini pesan teks
        if "message" in data and "text" in data["message"]:
            text = str(data["message"]["text"]).strip()
            chat_id = str(data["message"]["chat"]["id"])
            sender = data["message"]["from"].get("first_name", "User")

            # LOGIKA COMMAND /info
            if text == "/info" or text == "/start":
                print(f"📩 Menerima perintah '{text}' dari {sender}")
                
                # Kirim status "Sedang mengetik..." (Opsional)
                # Kita biarkan ini synchronous juga agar urut
                requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendChatAction", 
                              json={"chat_id": chat_id, "action": "typing"})
                
                # Generate Laporan
                report_msg = generate_upcoming_report_text()
                
                # 🔥 PERBAIKAN: KIRIM LANGSUNG (Direct Call)
                # Jangan pakai background_tasks di sini untuk Vercel Webhook
                send_telegram_notif(report_msg, chat_id)

        return {"status": "ok"}
    except Exception as e:
        print(f"Webhook Error: {e}")
        return {"status": "error"}

@app.get("/api/check-upcoming")
def manual_check_upcoming(background_tasks: BackgroundTasks):
    """Endpoint manual via Browser"""
    msg = generate_upcoming_report_text()
    # Untuk manual trigger via browser, background task biasanya OK, 
    # tapi agar aman di Vercel, kita direct call juga
    send_telegram_notif(msg) 
    return {"status": "Sent", "preview": msg}

@app.get("/api/analytics")
def get_analytics_data():
    try:
        response = supabase.table("letters").select("*").eq("is_deleted", False).execute()
        letters = cast(List[Dict[str, Any]], response.data or [])
        
        total_surat = len(letters)
        total_nominal = sum([int(float(str(l.get("nominal_jaminan",0)))) for l in letters if l.get("nominal_jaminan")])
        total_expired = len([l for l in letters if str(l.get("status")).lower() == "expired"])
        
        status_counts: Dict[str, int] = {}
        vendor_stats: Dict[str, int] = {}
        
        for item in letters:
            s = str(item.get("status", "Unknown"))
            status_counts[s] = status_counts.get(s, 0) + 1
            v = str(item.get("vendor", "Unknown"))
            n = int(float(str(item.get("nominal_jaminan",0)))) if item.get("nominal_jaminan") else 0
            vendor_stats[v] = vendor_stats.get(v, 0) + n

        pie_data = [{"name": k, "value": v, "color": "#10B981" if k=="Aktif" else "#EF4444"} for k, v in status_counts.items()]
        bar_data = [{"name": k[:15]+"...", "total": v} for k, v in sorted(vendor_stats.items(), key=lambda x:x[1], reverse=True)[:5]]

        return {"summary": {"total_surat": total_surat, "total_nominal": total_nominal, "total_expired": total_expired}, "pie_chart": pie_data, "bar_chart": bar_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/letters/active")
def get_active_letters():
    return supabase.table("letters").select("*").eq("is_deleted", False).neq("status", "Expired").neq("status", "Selesai").order("id", desc=True).execute().data or []

@app.get("/letters/archive")
def get_archived_letters():
    return supabase.table("letters").select("*").eq("is_deleted", False).or_("status.eq.Expired,status.eq.Selesai").order("id", desc=True).execute().data or []

@app.get("/letters")
def get_all_letters():
    return supabase.table("letters").select("*").eq("is_deleted", False).order("id", desc=True).execute().data or []

@app.get("/letters/{letter_id}")
def get_letter_by_id(letter_id: int):
    res = supabase.table("letters").select("*").eq("id", letter_id).single().execute()
    if not res.data: raise HTTPException(404)
    return res.data

@app.post("/letters")
def create_letter(letter: LetterSchema, background_tasks: BackgroundTasks):
    data = letter.dict(); user = data.pop("user_email", "Admin")
    if "id" in data: del data["id"]
    data["is_deleted"] = False
    data["vendor"] = sanitize_text(data["vendor"])
    data["pekerjaan"] = sanitize_text(data["pekerjaan"])
    res = supabase.table("letters").insert(data).execute()
    if res.data:
        background_tasks.add_task(log_activity_bg, user, "CREATE", f"Tambah: {data['vendor']}")
        msg = f"🆕 *DATA BARU*\n🏢 {data['vendor']}\n📄 `{data['nomor_kontrak']}`"
        # Untuk notifikasi create, background tasks biasanya OK karena user interaksi via frontend
        background_tasks.add_task(send_telegram_notif, msg)
    return res.data

@app.put("/letters/{letter_id}")
def update_letter(letter_id: int, letter: LetterSchema, background_tasks: BackgroundTasks):
    data = letter.dict(); user = data.pop("user_email", "Admin")
    if "id" in data: del data["id"]
    data["vendor"] = sanitize_text(data["vendor"])
    data["pekerjaan"] = sanitize_text(data["pekerjaan"])
    res = supabase.table("letters").update(data).eq("id", letter_id).execute()
    background_tasks.add_task(log_activity_bg, user, "UPDATE", f"Edit: {data['vendor']}")
    return {"status": "success"}

@app.delete("/letters/{letter_id}")
def delete_letter(letter_id: int, background_tasks: BackgroundTasks, user_email: str = "Admin"):
    # Fix Cast
    exist = supabase.table("letters").select("vendor").eq("id", letter_id).execute()
    d_list = cast(List[Dict[str, Any]], exist.data or [])
    target = str(d_list[0].get('vendor', 'Unknown')) if d_list else "Unknown"
    supabase.table("letters").update({"is_deleted": True}).eq("id", letter_id).execute()
    background_tasks.add_task(log_activity_bg, user_email, "SOFT_DELETE", f"Hapus: {target}")
    return {"status": "success"}

@app.get("/api/cron-update-status")
def cron_auto_update_status(background_tasks: BackgroundTasks):
    # Gunakan fungsi generate report yang sudah kita buat
    report = generate_upcoming_report_text()
    
    # Logic update expired database
    tz = pytz.timezone('Asia/Makassar'); today = datetime.now(tz).strftime('%Y-%m-%d')
    res = supabase.table("letters").select("id").eq("is_deleted", False).lt("tanggal_akhir_garansi", today).neq("status", "Expired").neq("status", "Selesai").execute()
    lst = cast(List[Dict[str, Any]], res.data or [])
    for x in lst:
        if x.get('id'): supabase.table("letters").update({"status": "Expired"}).eq("id", x['id']).execute()
    
    # Kirim report ke Default Group (Hanya saat pagi hari via Cron)
    # Cron Vercel punya timeout lebih panjang, jadi direct call lebih aman
    if "Sisa:" in report: 
        send_telegram_notif(report)
        
    background_tasks.add_task(log_activity_bg, "System", "AUTO_UPDATE", f"Check done. {len(lst)} updated.")
    
    return {"status": "success"}

@app.get("/export/excel")
def export_excel_multisheet():
    try:
        template_path = os.path.join(os.path.dirname(__file__), "Template_Sijagad.xlsx")
        
        if not os.path.exists(template_path):
             raise HTTPException(status_code=404, detail="Template tidak ditemukan.")

        wb = load_workbook(template_path)

        if "PELAKSANAAN" not in wb.sheetnames or "PEMELIHARAAN" not in wb.sheetnames:
            raise HTTPException(status_code=500, detail="Template salah format.")

        ws_pelaksanaan = wb["PELAKSANAAN"]
        ws_pemeliharaan = wb["PEMELIHARAAN"]

        response = supabase.table("letters").select("*").eq("is_deleted", False).order("id", desc=True).execute()
        data = response.data or []
        
        if not data:
             output = BytesIO()
             wb.save(output)
             output.seek(0)
             filename = f"Laporan_SiJAGAD_{datetime.now().strftime('%Y%m%d')}.xlsx"
             return StreamingResponse(output, headers={'Content-Disposition': f'attachment; filename="{filename}"'}, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        df = pd.DataFrame(data)
        df['kategori'] = df['kategori'].astype(str).str.strip()

        data_pelaksanaan = cast(List[Dict[str, Any]], df[df['kategori'].str.contains('Pelaksanaan', case=False, na=False)].to_dict('records'))
        data_pemeliharaan = cast(List[Dict[str, Any]], df[df['kategori'].str.contains('Pemeliharaan', case=False, na=False)].to_dict('records'))

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def fill_sheet(worksheet, data_list: List[Dict[str, Any]]):
            row_idx = 2 
            for i, item in enumerate(data_list, start=1):
                get_val = lambda key: str(item.get(key) if item.get(key) is not None else '-')
                get_nominal = lambda key: int(float(str(item.get(key, 0)))) if item.get(key) else 0

                worksheet.cell(row=row_idx, column=1, value=i)
                worksheet.cell(row=row_idx, column=2, value=get_val('vendor'))
                worksheet.cell(row=row_idx, column=3, value=get_val('pekerjaan'))
                worksheet.cell(row=row_idx, column=4, value=get_val('nomor_kontrak'))
                worksheet.cell(row=row_idx, column=5, value=get_val('tanggal_awal_kontrak'))
                
                cell_f = worksheet.cell(row=row_idx, column=6, value=get_nominal('nominal_jaminan'))
                cell_f.number_format = '#,##0' 
                
                worksheet.cell(row=row_idx, column=7, value=get_val('jenis_garansi'))
                worksheet.cell(row=row_idx, column=8, value=get_val('nomor_garansi'))
                worksheet.cell(row=row_idx, column=9, value=get_val('bank_penerbit'))
                worksheet.cell(row=row_idx, column=10, value=get_val('tanggal_awal_garansi'))
                worksheet.cell(row=row_idx, column=11, value=get_val('tanggal_akhir_garansi'))
                worksheet.cell(row=row_idx, column=12, value="")
                worksheet.cell(row=row_idx, column=13, value="") 

                for col_num in range(1, 14):
                    worksheet.cell(row=row_idx, column=col_num).border = thin_border
                
                row_idx += 1

        fill_sheet(ws_pelaksanaan, data_pelaksanaan)
        fill_sheet(ws_pemeliharaan, data_pemeliharaan)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Laporan_SiJAGAD_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Export Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal export: {str(e)}")

@app.get("/logs")
def get_logs():
    return supabase.table("activity_sijagad").select("*").order("created_at", desc=True).limit(50).execute().data or []